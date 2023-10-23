from openpyxl import load_workbook
path_to_xlsx_file = 'abinet_130.xlsx'


workbook = load_workbook(filename = path_to_xlsx_file)

sheet = workbook['Sheet1']

# get the maximun row and column in the sheet
max_row = sheet.max_row
max_column = sheet.max_column

for column in range(2, max_column + 1):
    column_value = []
    for row in range(1, max_row +1):
        cell_value = sheet.cell(row = row, column= column).value
        column_value.append(cell_value)
    
    